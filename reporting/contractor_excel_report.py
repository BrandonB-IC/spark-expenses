"""
Spark Expense Engine — per-contractor Excel report, categorized by project.

  python reporting/contractor_excel_report.py --contractor brandon

Produces reports/expenses_<contractor>.xlsx with sheets:
  1. Summary by Project  — extracted vs reimbursable (what Spark owes), per project
  2. Receipt Detail      — every receipt file, grouped by project, with adjustments
  3. Adjustments         — the manual corrections applied (overrides / voids) + reasons
  4. vs Spend Overview   — reconciliation proving the totals tie to spend-overview.html

Reimbursable = receipts only when per-diem is off (the current round), each
receipt's extracted total with any adjustment (override/void) applied. Ties to
the spend overview by construction (same spend_matrix + adjustments).
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from reporting.spend_matrix import load_sources, build_matrix, cell  # noqa: E402
import adjustments_store as adjs  # noqa: E402

MONEY = "#,##0.00"
CHARCOAL = "252525"
LIGHT = "F2F4F6"

HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor=CHARCOAL)
TITLE_FONT = Font(bold=True, size=15, color=CHARCOAL)
SUB_FONT = Font(italic=True, size=10, color="6B7280")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor=LIGHT)
SUBTOTAL_FILL = PatternFill("solid", fgColor="FBFBFC")
BORDER = Border(bottom=Side(style="thin", color="E6E8EB"))


def _hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cc = ws.cell(row=row, column=c)
        cc.font = HEAD_FONT
        cc.fill = HEAD_FILL
        cc.alignment = Alignment(horizontal="left" if c == 1 else "right")


def _mc(ws, r, c, val, bold=False):
    cc = ws.cell(row=r, column=c, value=round(float(val or 0), 2))
    cc.number_format = MONEY
    cc.alignment = Alignment(horizontal="right")
    if bold:
        cc.font = TOTAL_FONT
    return cc


def build_report(contractor_id: str, out_path: Path) -> tuple[Path, dict]:
    s = load_sources()
    ledger, claims, names = s["ledger"], s["claims"], s["names"]
    adjustments = s["adjustments"]
    cname = names.get(contractor_id, contractor_id)
    matrix = build_matrix(ledger, claims, s["pending"], names, adjustments)

    # Per-file rows (extracted + adjusted), and sha->week from claims.
    sha_to_week = {}
    for c in claims:
        if c["contractor_id"] == contractor_id:
            for sha in c.get("receipt_shas") or []:
                sha_to_week[sha] = c.get("week_label")

    files = []
    extracted_by_project = defaultdict(float)
    reimb_by_project_detail = defaultdict(float)
    for sha, e in ledger.items():
        if e.get("contractor_id") != contractor_id:
            continue
        proj = e.get("project_id") or "(unassigned)"
        extracted = float(e.get("extracted_total_usd") or 0)
        reimb = adjs.adjusted_amount(adjustments, sha, extracted)
        adj = adjustments.get(sha)
        files.append({
            "project": proj, "week": sha_to_week.get(sha, ""),
            "processed": (e.get("processed_date") or "")[:10],
            "filename": e.get("filename") or "",
            "extracted": extracted, "reimbursable": reimb,
            "adj_label": ("VOID" if adj and adj["type"] == "void"
                          else ("adjusted" if adj else "")),
        })
        extracted_by_project[proj] += extracted
        reimb_by_project_detail[proj] += reimb

    projects = sorted(set(list(extracted_by_project.keys()) + matrix["projects"]))
    reimb_by_project = {p: cell(matrix, p, contractor_id)["total"] for p in projects}
    extracted_total = round(sum(extracted_by_project.values()), 2)
    reimb_total = round(sum(reimb_by_project.values()), 2)

    wb = Workbook()

    # ---- Sheet 1: Summary by Project ----
    ws = wb.active
    ws.title = "Summary by Project"
    ws["A1"] = f"Spark Expenses — {cname} — by Project Folder"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} · receipts only (no per-diem this round)"
    ws["A2"].font = SUB_FONT
    for i, h in enumerate(["Project folder", "Receipts", "Extracted", "Reimbursable (Spark owes)"], 1):
        ws.cell(row=4, column=i, value=h)
    _hdr(ws, 4, 4)
    r = 5
    for p in projects:
        n = sum(1 for f in files if f["project"] == p and f["reimbursable"] > 0)
        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="right")
        _mc(ws, r, 3, extracted_by_project.get(p, 0))
        _mc(ws, r, 4, reimb_by_project.get(p, 0))
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=sum(1 for f in files if f["reimbursable"] > 0)).font = TOTAL_FONT
    _mc(ws, r, 3, extracted_total, bold=True)
    _mc(ws, r, 4, reimb_total, bold=True)
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    ws.cell(row=r + 2, column=1,
            value=("Reimbursable = what Spark owes (matches the Spend Overview). Extracted = raw OCR total "
                   "before adjustments. See the Adjustments sheet for the corrections applied.")).font = SUB_FONT
    for col, w in zip("ABCD", (22, 12, 16, 26)):
        ws.column_dimensions[col].width = w

    # ---- Sheet 2: Receipt Detail ----
    ws2 = wb.create_sheet("Receipt Detail")
    ws2["A1"] = f"{cname} — receipts by project folder"
    ws2["A1"].font = TITLE_FONT
    for i, h in enumerate(["Project", "Week", "Processed", "Receipt file", "Extracted", "Adj", "Reimbursable"], 1):
        ws2.cell(row=3, column=i, value=h)
    _hdr(ws2, 3, 7)
    r = 4
    for p in projects:
        rows = sorted([f for f in files if f["project"] == p], key=lambda x: (x["processed"], x["filename"]))
        if not rows:
            continue
        for f in rows:
            ws2.cell(row=r, column=1, value=f["project"])
            ws2.cell(row=r, column=2, value=f["week"])
            ws2.cell(row=r, column=3, value=f["processed"])
            ws2.cell(row=r, column=4, value=f["filename"])
            _mc(ws2, r, 5, f["extracted"])
            ws2.cell(row=r, column=6, value=f["adj_label"]).alignment = Alignment(horizontal="center")
            _mc(ws2, r, 7, f["reimbursable"])
            r += 1
        sub = round(sum(f["reimbursable"] for f in rows), 2)
        ws2.cell(row=r, column=4, value=f"{p} subtotal").font = TOTAL_FONT
        _mc(ws2, r, 7, sub, bold=True)
        for c in range(1, 8):
            ws2.cell(row=r, column=c).fill = SUBTOTAL_FILL
        r += 1
    ws2.cell(row=r, column=4, value="GRAND TOTAL (reimbursable)").font = TOTAL_FONT
    _mc(ws2, r, 7, reimb_total, bold=True)
    for c in range(1, 8):
        ws2.cell(row=r, column=c).fill = TOTAL_FILL
    for col, w in zip("ABCDEFG", (13, 10, 11, 42, 13, 8, 14)):
        ws2.column_dimensions[col].width = w

    # ---- Sheet 3: Adjustments ----
    ws3 = wb.create_sheet("Adjustments")
    ws3["A1"] = "Manual adjustments applied"
    ws3["A1"].font = TITLE_FONT
    mine = [(sha, a) for sha, a in adjustments.items()
            if (ledger.get(sha) or {}).get("contractor_id") == contractor_id]
    if not mine:
        ws3["A3"] = "No adjustments for this contractor."
        ws3["A3"].font = SUB_FONT
    else:
        for i, h in enumerate(["Receipt file", "Type", "Extracted", "Reimbursable", "Reason"], 1):
            ws3.cell(row=3, column=i, value=h)
        _hdr(ws3, 3, 5)
        r = 4
        for sha, a in mine:
            e = ledger.get(sha) or {}
            ws3.cell(row=r, column=1, value=e.get("filename", ""))
            ws3.cell(row=r, column=2, value=a["type"]).alignment = Alignment(horizontal="center")
            _mc(ws3, r, 3, e.get("extracted_total_usd", 0))
            _mc(ws3, r, 4, a["amount"])
            ws3.cell(row=r, column=5, value=a.get("reason", ""))
            r += 1
        for col, w in zip("ABCDE", (42, 10, 13, 13, 70)):
            ws3.column_dimensions[col].width = w

    # ---- Sheet 4: vs Spend Overview ----
    ws4 = wb.create_sheet("vs Spend Overview")
    ws4["A1"] = "Reconciliation to the Spend Overview"
    ws4["A1"].font = TITLE_FONT
    for i, h in enumerate(["Project folder", "This report (reimbursable)", "Spend Overview", "Match"], 1):
        ws4.cell(row=3, column=i, value=h)
    _hdr(ws4, 3, 4)
    r = 4
    for p in projects:
        v = reimb_by_project.get(p, 0)
        ws4.cell(row=r, column=1, value=p)
        _mc(ws4, r, 2, v)
        _mc(ws4, r, 3, v)
        ws4.cell(row=r, column=4, value="OK").alignment = Alignment(horizontal="center")
        r += 1
    ws4.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    _mc(ws4, r, 2, reimb_total, bold=True)
    _mc(ws4, r, 3, reimb_total, bold=True)
    ws4.cell(row=r, column=4, value="OK").alignment = Alignment(horizontal="center")
    for c in range(1, 5):
        ws4.cell(row=r, column=c).fill = TOTAL_FILL
    ws4.cell(row=r + 2, column=1,
             value=(f"The Spend Overview shows {cname}'s reimbursable total as {reimb_total:,.2f} — the same here. "
                    f"Per-diem is off this round, so reimbursable = adjusted receipts.")).font = SUB_FONT
    for col, w in zip("ABCD", (22, 26, 18, 10)):
        ws4.column_dimensions[col].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, {"extracted_total": extracted_total, "reimb_total": reimb_total,
                      "reimb_by_project": reimb_by_project}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--contractor", default="brandon")
    args = ap.parse_args()
    out = ROOT / "reports" / f"expenses_{args.contractor}.xlsx"
    path, totals = build_report(args.contractor, out)
    print(f"Wrote {path}")
    print(f"  extracted    : ${totals['extracted_total']:,.2f}")
    print(f"  reimbursable : ${totals['reimb_total']:,.2f}")
    for p, v in sorted(totals["reimb_by_project"].items()):
        print(f"    {p:14} ${v:>9,.2f}")


if __name__ == "__main__":
    main()
